"""
Real-Company Ratio Comparison Tool — for student learning
------------------------------------------------------------
Pulls real financial statements for up to 4 public companies via
yfinance, computes 17 fundamental ratios PLUS 4 market/valuation
ratios (P/E, P/S, P/B, Book-to-Market) across two fiscal years you
choose, compares companies side-by-side, and benchmarks them against
illustrative industry-average ratios.
Year selection is a dropdown, not a hardcoded year — the app always
fetches whatever annual periods Yahoo Finance currently has for each
company, and you pick which two to compare. That means next year (or
in five years) you can keep using this same file: fetch again, pick
the newer years from the dropdown, done. No code changes needed.
IMPORTANT — data quality note for students:
Yahoo Finance's data (via the free yfinance library) is convenient
but not always complete or perfectly labeled — some companies are
missing a line item, or Yahoo reports it under an unexpected name.
That's why every fetched number is shown in an EDITABLE table before
ratios are computed: always sanity-check the auto-filled figures
against the company's actual 10-K before drawing conclusions.
AI INTERPRETATION (bring-your-own-chat): the Summary Dashboard tab
builds a ready-to-copy prompt from the already-computed ratio table and
gives students direct links to open their own ChatGPT / Gemini / Claude /
Perplexity tab (using whatever account they're already logged into — no
API key, no shared credentials, nothing this app touches). The student
pastes the prompt in, gets an answer, and pastes the response back into
this app to display alongside the numbers and include in the exported
report.
Run with:  streamlit run app.py
"""
import os
import io
import json
import time
from datetime import datetime
from xml.sax.saxutils import escape as xml_escape
import pandas as pd
import streamlit as st
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter as PAGE_SIZE
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
import yaml
import streamlit_authenticator as stauth
st.set_page_config(page_title="Real-Company Ratio Comparison Tool", page_icon="🏢", layout="wide")
MAX_PERIODS = 4  # how many annual periods to fetch per company (lets you pick older years too)
# ============================================================
# GENERIC HELPERS
# ============================================================
def safe_div(numerator, denominator):
    if numerator is None or denominator in (0, None):
        return None
    return numerator / denominator
def fmt(value, suffix=""):
    if value is None:
        return "N/A"
    if suffix == "%":
        return f"{value * 100:,.1f}%"
    if suffix == " days":
        return f"{value:,.0f} days"
    if suffix == "x":
        return f"{value:,.2f}x"
    return f"{value:,.2f}"
def fmt_delta(delta, suffix):
    if delta is None:
        return None
    if suffix == "%":
        return f"{delta * 100:+.1f} pp"
    if suffix == " days":
        return f"{delta:+.0f} days"
    if suffix == "x":
        return f"{delta:+.2f}x"
    return f"{delta:+.2f}"
def display_magnitude(value, suffix):
    """Converts a raw ratio value into the same units shown by fmt() /
    asked of the AI in the Calc Challenge prompt (e.g. 0.245 -> 24.5 for
    percentages), so a student-entered number can be compared apples-to-
    apples against the app's own computed value."""
    if value is None:
        return None
    if suffix == "%":
        return value * 100
    return value
def compare_flag(a, b):
    """5%-relative-tolerance match flag for comparing two ratio values
    that are already in the same display units (see display_magnitude).
    Used throughout the AI Calc Challenge tab's comparison tables."""
    if a is None or b is None:
        return "⚪ N/A"
    tolerance = max(abs(a) * 0.05, 0.05)
    return "✅ Match" if abs(a - b) <= tolerance else "❌ Differ"

MAX_TEAM_MEMBERS = 4
def compute_prepared_by(team_members_raw, fallback_name):
    """Turns the sidebar's free-text 'Team members' field into the string
    stamped as 'Prepared by:' on every export (Excel/PDF). Falls back to
    just the logged-in student's name if the field is left blank, and caps
    at MAX_TEAM_MEMBERS names (extras are dropped, not silently merged, so
    the caller can warn the student). Returns (display_string, was_truncated).
    """
    names = [n.strip() for n in (team_members_raw or "").split(",") if n.strip()]
    if not names:
        names = [fallback_name]
    truncated = len(names) > MAX_TEAM_MEMBERS
    return ", ".join(names[:MAX_TEAM_MEMBERS]), truncated

# Remembers each student's last-typed "Team members" text between logins, so
# they don't have to retype teammates' names every session. Deliberately
# lightweight — a small local JSON file, not the Google Sheets progress
# store used by the AI Calc Challenge tab. That means: on your own machine
# or a self-hosted server this just works, persisting indefinitely. On
# Streamlit Community Cloud it survives ordinary logins but is NOT
# guaranteed to survive a redeploy (a new git push) or occasional platform
# container recycling, since that disk isn't a persistent volume — an
# acceptable tradeoff here since losing this just means retyping team
# names once, unlike the AI Calc Challenge's graded answers.
TEAM_MEMBERS_FILE = "team_members.json"
def load_team_members(username):
    """Returns the last-saved 'Team members' text for this student, or ''
    if nothing's saved yet (including if the file doesn't exist)."""
    try:
        with open(TEAM_MEMBERS_FILE) as f:
            data = json.load(f)
        return data.get(username, "")
    except Exception:
        return ""

def save_team_members(username, value):
    """Best-effort save — never lets a filesystem hiccup crash the app."""
    try:
        try:
            with open(TEAM_MEMBERS_FILE) as f:
                data = json.load(f)
        except Exception:
            data = {}
        data[username] = value
        with open(TEAM_MEMBERS_FILE, "w") as f:
            json.dump(data, f)
    except Exception:
        pass
AI_TOOL_OPTIONS = ["ChatGPT", "Gemini", "Claude", "Perplexity", "Other"]
def record_ai_disclosure(prefix, ticker, response_text, saved_tool=None, saved_timestamp=None):
    """Renders a 'which AI did you use' selector next to a pasted AI
    response, and silently timestamps the moment that pasted text last
    changed. This turns the disclosure record into structured metadata
    (tool + when) instead of relying on whatever the AI happened to
    mention about itself in its own reply. Returns (tool_name, timestamp)
    — timestamp is None until some response text has been pasted.

    saved_tool/saved_timestamp let a previously-autosaved disclosure (see
    the PERSISTENT PROGRESS section) seed this widget the first time it's
    shown in a session, so a returning student sees their prior tool/time
    instead of the exercise looking un-started."""
    tool_key = f"{prefix}_tool_{ticker}"
    stamp_key = f"{prefix}_stamp_{ticker}"
    seen_key = f"{prefix}_seen_{ticker}"
    if saved_tool and tool_key not in st.session_state:
        st.session_state[tool_key] = saved_tool
    if saved_timestamp and stamp_key not in st.session_state:
        st.session_state[stamp_key] = saved_timestamp
        st.session_state.setdefault(seen_key, response_text)
    tool = st.selectbox("Which AI did you use?", AI_TOOL_OPTIONS, key=tool_key)
    if response_text and response_text.strip():
        if st.session_state.get(seen_key) != response_text:
            st.session_state[stamp_key] = datetime.now().strftime("%Y-%m-%d %H:%M")
            st.session_state[seen_key] = response_text
    timestamp = st.session_state.get(stamp_key)
    if timestamp:
        st.caption(f"🕒 Recorded: {timestamp} (local time when you pasted/last edited this response) · Tool: {tool}")
    return tool, timestamp
# ============================================================
# PERSISTENT PROGRESS (optional — Google Sheets)
# ============================================================
# The AI Calc Challenge tab autosaves each student's in-progress answers so
# they survive closing the tab and logging back in later, including on
# Streamlit Community Cloud where local disk isn't reliably persistent
# across redeploys. This is entirely optional: if the instructor hasn't
# configured a Google Sheet (see README.md → "Persistent progress"), every
# function below quietly returns "no store" and the app falls back to
# ordinary in-browser-session-only behavior.
PROGRESS_HEADER = ["username", "ticker", "y2_label", "updated_at", "data_json"]

@st.cache_resource(show_spinner=False)
def _progress_worksheet():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        return None
    if "gcp_service_account" not in st.secrets or "progress_sheet_id" not in st.secrets:
        return None
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=scopes
        )
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(st.secrets["progress_sheet_id"])
        try:
            ws = sh.worksheet("progress")
        except Exception:
            ws = sh.add_worksheet(title="progress", rows=2000, cols=len(PROGRESS_HEADER))
            ws.append_row(PROGRESS_HEADER)
        return ws
    except Exception:
        return None

def progress_store_available():
    return _progress_worksheet() is not None

def _progress_find_row(ws, username, ticker, y2_label):
    try:
        values = ws.get_all_values()
    except Exception:
        return None, []
    for i, row in enumerate(values[1:], start=2):
        if len(row) >= 3 and row[0] == username and row[1] == ticker and row[2] == y2_label:
            return i, values
    return None, values

def load_progress(username, ticker, y2_label):
    """Returns the last-autosaved answers dict for this student+company, or
    {} if nothing is saved yet (or no store is configured)."""
    ws = _progress_worksheet()
    if ws is None:
        return {}
    idx, values = _progress_find_row(ws, username, ticker, y2_label)
    if idx is None:
        return {}
    row = values[idx - 1]
    if len(row) >= 5 and row[4]:
        try:
            return json.loads(row[4])
        except Exception:
            return {}
    return {}

def save_progress(username, ticker, y2_label, data):
    ws = _progress_worksheet()
    if ws is None:
        return False
    try:
        idx, _ = _progress_find_row(ws, username, ticker, y2_label)
        payload = [username, ticker, y2_label,
                   datetime.now().strftime("%Y-%m-%d %H:%M:%S"), json.dumps(data)]
        if idx:
            ws.update(f"A{idx}:E{idx}", [payload])
        else:
            ws.append_row(payload)
        return True
    except Exception:
        return False

def autosave_progress(username, ticker, y2_label, data):
    """Writes `data` only if it changed since the last autosave this
    session (avoids hammering the Sheets API on every rerun), and records
    the outcome in session_state so the UI can show a small status line."""
    hash_key = f"_prog_hash_{ticker}_{y2_label}"
    fresh_hash = hash(json.dumps(data, sort_keys=True, default=str))
    if st.session_state.get(hash_key) == fresh_hash:
        return
    ok = save_progress(username, ticker, y2_label, data)
    st.session_state[hash_key] = fresh_hash if ok else None
    st.session_state[f"_prog_saved_at_{ticker}_{y2_label}"] = (
        datetime.now().strftime("%H:%M:%S") if ok else None
    )

def load_progress_for_company(username, ticker, y2_label):
    """Loads saved progress once per browser session (guarded by
    session_state so we don't re-hit the Sheets API on every rerun), and
    seeds the one plain (non-widget-default) flag — the Round 2 prediction
    lock — that can't be seeded via a widget's `value=` argument."""
    guard_key = f"_prog_loaded_{ticker}_{y2_label}"
    if guard_key in st.session_state:
        return st.session_state.get(f"_prog_data_{ticker}_{y2_label}", {})
    data = load_progress(username, ticker, y2_label)
    st.session_state[guard_key] = True
    st.session_state[f"_prog_data_{ticker}_{y2_label}"] = data
    if data.get("predict_confirmed"):
        st.session_state.setdefault(f"predict_confirmed_{ticker}", True)
    return data
# ============================================================
# INTERPRETATION — fundamental ratios get Strong/Moderate/Weak badges.
# Market valuation ratios get a separate Low/Average/High "pricing"
# tier, since a low P/E isn't inherently "better" the way a strong
# current ratio is — it can just mean lower growth expectations.
# ============================================================
def _tier(v, good, ok):
    if v is None:
        return "na"
    if good(v):
        return "good"
    if ok(v):
        return "ok"
    return "warning"
def level_current_ratio(v):   return _tier(v, lambda x: x >= 2, lambda x: x >= 1)
def level_quick_ratio(v):     return _tier(v, lambda x: x >= 1, lambda x: x >= 0.7)
def level_cash_ratio(v):      return _tier(v, lambda x: x >= 0.5, lambda x: x >= 0.2)
def level_gross_margin(v):    return _tier(v, lambda x: x >= 0.40, lambda x: x >= 0.20)
def level_operating_margin(v):return _tier(v, lambda x: x >= 0.15, lambda x: x >= 0.05)
def level_net_margin(v):      return _tier(v, lambda x: x >= 0.10, lambda x: x >= 0.02)
def level_roa(v):             return _tier(v, lambda x: x >= 0.10, lambda x: x >= 0.03)
def level_roe(v):             return _tier(v, lambda x: x >= 0.15, lambda x: x >= 0.05)
def level_asset_turnover(v):  return _tier(v, lambda x: x >= 1.0, lambda x: x >= 0.5)
def level_inventory_turnover(v): return _tier(v, lambda x: x >= 8, lambda x: x >= 3)
def level_dio(v):              return "na" if v is None else ("good" if v <= 45 else ("ok" if v <= 90 else "warning"))
def level_receivables_turnover(v): return _tier(v, lambda x: x >= 10, lambda x: x >= 4)
def level_dso(v):               return "na" if v is None else ("good" if v <= 45 else ("ok" if v <= 90 else "warning"))
def level_debt_ratio(v):        return "na" if v is None else ("good" if v <= 0.4 else ("ok" if v <= 0.6 else "warning"))
def level_debt_to_equity(v):    return "na" if v is None else ("good" if v <= 1.0 else ("ok" if v <= 2.0 else "warning"))
def level_equity_multiplier(v): return "na" if v is None else ("good" if v <= 2.0 else ("ok" if v <= 3.0 else "warning"))
def level_interest_coverage(v): return _tier(v, lambda x: x >= 5, lambda x: x >= 1.5)
def level_pe(v):
    if v is None: return "na"
    if v < 15: return "low"
    if v <= 25: return "avg"
    return "high"
def level_ps(v):
    if v is None: return "na"
    if v < 1: return "low"
    if v <= 4: return "avg"
    return "high"
def level_pb(v):
    if v is None: return "na"
    if v < 1: return "low"
    if v <= 3: return "avg"
    return "high"
def level_btm(v):
    if v is None: return "na"
    if v >= 1: return "low"    # high book-to-market = cheap = "low" pricing tier
    if v >= 0.3: return "avg"
    return "high"
LEVEL_LABEL = {"good": "✅ Strong", "ok": "🟡 Moderate", "warning": "🔴 Weak", "na": "⚪ N/A"}
LEVEL_LABEL_VALUATION = {"low": "🟢 Low (cheaper)", "avg": "🟡 Average", "high": "🔴 High (pricier)", "na": "⚪ N/A"}
# ============================================================
# RATIO DEFINITIONS (single source of truth)
# ============================================================
RATIO_DEFS = [
    dict(key="current_ratio", category="Liquidity", name="Current Ratio",
         formula="Current Assets ÷ Current Liabilities", suffix="", level_fn=level_current_ratio,
         guide="Rule of thumb: ≥2 strong · 1–2 moderate · <1 weak (current liabilities exceed current assets)."),
    dict(key="quick_ratio", category="Liquidity", name="Quick Ratio (Acid-Test)",
         formula="(Current Assets − Inventory) ÷ Current Liabilities", suffix="", level_fn=level_quick_ratio,
         guide="Rule of thumb: ≥1 strong · 0.7–1 moderate · <0.7 weak. Excludes inventory from liquid assets."),
    dict(key="cash_ratio", category="Liquidity", name="Cash Ratio",
         formula="Cash ÷ Current Liabilities", suffix="", level_fn=level_cash_ratio,
         guide="Rule of thumb: ≥0.5 strong · 0.2–0.5 moderate · <0.2 weak. The most conservative liquidity measure."),
    dict(key="gross_margin", category="Profitability", name="Gross Profit Margin",
         formula="Gross Profit ÷ Sales", suffix="%", level_fn=level_gross_margin,
         guide="Rule of thumb: ≥40% strong · 20–40% moderate · <20% weak. Varies enormously by industry."),
    dict(key="operating_margin", category="Profitability", name="Operating Profit Margin",
         formula="EBIT ÷ Sales", suffix="%", level_fn=level_operating_margin,
         guide="Rule of thumb: ≥15% strong · 5–15% moderate · <5% weak."),
    dict(key="net_margin", category="Profitability", name="Net Profit Margin",
         formula="Net Income ÷ Sales", suffix="%", level_fn=level_net_margin,
         guide="Rule of thumb: ≥10% strong · 2–10% moderate · <2% weak."),
    dict(key="roa", category="Profitability", name="Return on Assets (ROA)",
         formula="Net Income ÷ Total Assets", suffix="%", level_fn=level_roa,
         guide="Rule of thumb: ≥10% strong · 3–10% moderate · <3% weak."),
    dict(key="roe", category="Profitability", name="Return on Equity (ROE)",
         formula="Net Income ÷ Shareholders' Equity", suffix="%", level_fn=level_roe,
         guide="Rule of thumb: ≥15% strong · 5–15% moderate · <5% weak."),
    dict(key="asset_turnover", category="Efficiency", name="Asset Turnover",
         formula="Sales ÷ Total Assets", suffix="x", level_fn=level_asset_turnover,
         guide="Rule of thumb: ≥1.0x strong · 0.5–1.0x moderate · <0.5x weak. Capital-intensive firms run lower."),
    dict(key="inventory_turnover", category="Efficiency", name="Inventory Turnover",
         formula="COGS ÷ Inventory", suffix="x", level_fn=level_inventory_turnover,
         guide="Rule of thumb: ≥8x strong · 3–8x moderate · <3x weak."),
    dict(key="dio", category="Efficiency", name="Days Inventory Outstanding",
         formula="365 ÷ Inventory Turnover", suffix=" days", level_fn=level_dio,
         guide="Rule of thumb: ≤45 days strong · 45–90 moderate · >90 weak (lower is better)."),
    dict(key="receivables_turnover", category="Efficiency", name="Receivables Turnover",
         formula="Sales ÷ Accounts Receivable", suffix="x", level_fn=level_receivables_turnover,
         guide="Rule of thumb: ≥10x strong · 4–10x moderate · <4x weak."),
    dict(key="dso", category="Efficiency", name="Days Sales Outstanding (DSO)",
         formula="365 ÷ Receivables Turnover", suffix=" days", level_fn=level_dso,
         guide="Rule of thumb: ≤45 days strong · 45–90 moderate · >90 weak (lower is better)."),
    dict(key="debt_ratio", category="Solvency", name="Debt Ratio",
         formula="Total Liabilities ÷ Total Assets", suffix="%", level_fn=level_debt_ratio,
         guide="Rule of thumb: ≤40% strong · 40–60% moderate · >60% weak (lower is better)."),
    dict(key="debt_to_equity", category="Solvency", name="Debt-to-Equity Ratio",
         formula="Total Liabilities ÷ Equity", suffix="x", level_fn=level_debt_to_equity,
         guide="Rule of thumb: ≤1.0x strong · 1–2x moderate · >2x weak (lower is better)."),
    dict(key="equity_multiplier", category="Solvency", name="Equity Multiplier",
         formula="Total Assets ÷ Equity", suffix="x", level_fn=level_equity_multiplier,
         guide="Rule of thumb: ≤2.0x strong · 2–3x moderate · >3x weak (lower is better)."),
    dict(key="interest_coverage", category="Solvency", name="Interest Coverage Ratio",
         formula="EBIT ÷ Interest Expense", suffix="x", level_fn=level_interest_coverage,
         guide="Rule of thumb: ≥5x strong · 1.5–5x moderate · <1.5x weak."),
    dict(key="pe_ratio", category="Market Valuation", name="Price/Earnings (P/E)",
         formula="Market Cap ÷ Net Income", suffix="x", level_fn=level_pe,
         guide="Rough tiers: <15x low · 15–25x average · >25x high. Low isn't automatically 'better' — it can reflect lower expected growth or higher perceived risk."),
    dict(key="ps_ratio", category="Market Valuation", name="Price/Sales (P/S)",
         formula="Market Cap ÷ Revenue", suffix="x", level_fn=level_ps,
         guide="Rough tiers: <1x low · 1–4x average · >4x high. Useful when a company isn't yet profitable and P/E doesn't apply."),
    dict(key="pb_ratio", category="Market Valuation", name="Price/Book (P/B)",
         formula="Market Cap ÷ Book Value of Equity", suffix="x", level_fn=level_pb,
         guide="Rough tiers: <1x low · 1–3x average · >3x high. Below 1x means the market values the company below its accounting book value."),
    dict(key="book_to_market", category="Market Valuation", name="Book-to-Market Ratio",
         formula="Book Value of Equity ÷ Market Cap", suffix="x", level_fn=level_btm,
         guide="The inverse of P/B. Higher = more 'value'-like (cheap relative to book); lower = more 'growth'-like (market pays a premium over book)."),
]
CATEGORIES = ["Liquidity", "Profitability", "Efficiency", "Solvency", "Market Valuation"]
# ============================================================
# DERIVED FIGURES & RATIO CALCULATION
# ============================================================
def compute_derived(v):
    v = dict(v)
    if v.get("current_assets") is None and all(v.get(k) is not None for k in ("cash", "receivables", "inventory")):
        v["current_assets"] = v["cash"] + v["receivables"] + v["inventory"]
    if v.get("total_liabilities") is None and v.get("total_assets") is not None and v.get("equity") is not None:
        v["total_liabilities"] = v["total_assets"] - v["equity"]
    if v.get("equity") is None and v.get("total_assets") is not None and v.get("total_liabilities") is not None:
        v["equity"] = v["total_assets"] - v["total_liabilities"]
    if v.get("gross_profit") is None and v.get("revenue") is not None and v.get("cogs") is not None:
        v["gross_profit"] = v["revenue"] - v["cogs"]
    return v
def compute_ratios(d, market_cap=None):
    r = {}
    r["current_ratio"] = safe_div(d.get("current_assets"), d.get("current_liabilities"))
    r["quick_ratio"] = safe_div(
        None if d.get("current_assets") is None or d.get("inventory") is None else d["current_assets"] - d["inventory"],
        d.get("current_liabilities"),
    )
    r["cash_ratio"] = safe_div(d.get("cash"), d.get("current_liabilities"))
    r["gross_margin"] = safe_div(d.get("gross_profit"), d.get("revenue"))
    r["operating_margin"] = safe_div(d.get("ebit"), d.get("revenue"))
    r["net_margin"] = safe_div(d.get("net_income"), d.get("revenue"))
    r["roa"] = safe_div(d.get("net_income"), d.get("total_assets"))
    r["roe"] = safe_div(d.get("net_income"), d.get("equity"))
    r["asset_turnover"] = safe_div(d.get("revenue"), d.get("total_assets"))
    r["inventory_turnover"] = safe_div(d.get("cogs"), d.get("inventory"))
    r["dio"] = safe_div(365, r["inventory_turnover"]) if r["inventory_turnover"] else None
    r["receivables_turnover"] = safe_div(d.get("revenue"), d.get("receivables"))
    r["dso"] = safe_div(365, r["receivables_turnover"]) if r["receivables_turnover"] else None
    r["debt_ratio"] = safe_div(d.get("total_liabilities"), d.get("total_assets"))
    r["debt_to_equity"] = safe_div(d.get("total_liabilities"), d.get("equity"))
    r["equity_multiplier"] = safe_div(d.get("total_assets"), d.get("equity"))
    r["interest_coverage"] = safe_div(d.get("ebit"), d.get("interest_expense"))
    r["pe_ratio"] = safe_div(market_cap, d.get("net_income"))
    r["ps_ratio"] = safe_div(market_cap, d.get("revenue"))
    r["pb_ratio"] = safe_div(market_cap, d.get("equity"))
    r["book_to_market"] = safe_div(d.get("equity"), market_cap)
    return r
# ============================================================
# INDUSTRY BENCHMARKS
# 15 of 21 ratios below are REAL sector-aggregate data — current ratio,
# quick ratio, gross/operating/net margin, ROA, ROE, P/E, P/S, P/B are
# sourced directly from Finviz.com sector groups (as of Aug 18, 2026).
# Equity multiplier, debt ratio, debt-to-equity, asset turnover, and
# book-to-market are DERIVED from those real ROE/ROA/P-B figures via
# the DuPont identity (ROE = Net Margin x Asset Turnover x Equity
# Multiplier) — legitimate math on real data, not invented numbers.
# The remaining 6 ratios (cash ratio, inventory turnover, DIO,
# receivables turnover, DSO, interest coverage) have no reliable free
# sector-level source and stay as illustrative, instructor-editable
# estimates — see the Source column in the benchmark table below.
# This is a DATED SNAPSHOT, not a live feed — refresh periodically by
# checking finviz.com/groups.ashx?g=sector&v=160 (ROA/ROE/margins/
# ratios) and v=120 (P/E, P/S, P/B) and updating the numbers below.
# ============================================================
REAL_RATIO_KEYS = {
    "current_ratio", "quick_ratio", "gross_margin", "operating_margin", "net_margin",
    "roa", "roe", "asset_turnover", "debt_ratio", "debt_to_equity", "equity_multiplier",
    "pe_ratio", "ps_ratio", "pb_ratio", "book_to_market",
}
BENCHMARK_AS_OF = "August 18, 2026"
BENCHMARKS = {
    "Technology": dict(current_ratio=2.27, quick_ratio=1.98, cash_ratio=1.0, gross_margin=0.4910, operating_margin=0.2732,
                        net_margin=0.2220, roa=0.1334, roe=0.2889, asset_turnover=0.60, inventory_turnover=8, dio=45,
                        receivables_turnover=8, dso=45, debt_ratio=0.5382, debt_to_equity=1.17, equity_multiplier=2.17,
                        interest_coverage=15, pe_ratio=34.45, ps_ratio=8.25, pb_ratio=11.0, book_to_market=0.0909),
    "Healthcare": dict(current_ratio=2.13, quick_ratio=1.88, cash_ratio=0.5, gross_margin=0.3517, operating_margin=0.0958,
                        net_margin=0.0466, roa=0.0389, roe=0.0980, asset_turnover=0.83, inventory_turnover=5, dio=73,
                        receivables_turnover=7, dso=52, debt_ratio=0.6031, debt_to_equity=1.52, equity_multiplier=2.52,
                        interest_coverage=8, pe_ratio=32.50, ps_ratio=1.97, pb_ratio=4.82, book_to_market=0.2075),
    "Financial Services": dict(current_ratio=2.20, quick_ratio=3.32, cash_ratio=None, gross_margin=0.4077,
                                operating_margin=0.1972, net_margin=0.1481, roa=0.0129, roe=0.1261, asset_turnover=0.09,
                                inventory_turnover=None, dio=None, receivables_turnover=None, dso=None,
                                debt_ratio=0.8977, debt_to_equity=8.78, equity_multiplier=9.78, interest_coverage=2.5,
                                pe_ratio=16.98, ps_ratio=2.37, pb_ratio=2.26, book_to_market=0.4425),
    "Consumer Cyclical": dict(current_ratio=1.43, quick_ratio=1.05, cash_ratio=0.2, gross_margin=0.2934, operating_margin=0.0755,
                               net_margin=0.0601, roa=0.0510, roe=0.1260, asset_turnover=0.85, inventory_turnover=6, dio=61,
                               receivables_turnover=20, dso=18, debt_ratio=0.5952, debt_to_equity=1.47, equity_multiplier=2.47,
                               interest_coverage=6, pe_ratio=25.56, ps_ratio=1.88, pb_ratio=4.28, book_to_market=0.2336),
    "Consumer Defensive": dict(current_ratio=1.07, quick_ratio=0.66, cash_ratio=0.15, gross_margin=0.2819, operating_margin=0.0904,
                                net_margin=0.0516, roa=0.0608, roe=0.1589, asset_turnover=1.18, inventory_turnover=8, dio=46,
                                receivables_turnover=12, dso=30, debt_ratio=0.6174, debt_to_equity=1.61, equity_multiplier=2.61,
                                interest_coverage=8, pe_ratio=26.16, ps_ratio=1.40, pb_ratio=5.11, book_to_market=0.1957),
    "Industrials": dict(current_ratio=2.29, quick_ratio=1.94, cash_ratio=0.3, gross_margin=0.2484, operating_margin=0.1144,
                         net_margin=0.0763, roa=0.0539, roe=0.1619, asset_turnover=0.71, inventory_turnover=6, dio=61,
                         receivables_turnover=8, dso=46, debt_ratio=0.6671, debt_to_equity=2.00, equity_multiplier=3.00,
                         interest_coverage=7, pe_ratio=40.58, ps_ratio=3.35, pb_ratio=6.93, book_to_market=0.1443),
    "Energy": dict(current_ratio=1.34, quick_ratio=1.09, cash_ratio=0.3, gross_margin=0.2046, operating_margin=0.1515,
                    net_margin=0.0923, roa=0.0621, roe=0.1479, asset_turnover=0.67, inventory_turnover=10, dio=37,
                    receivables_turnover=10, dso=37, debt_ratio=0.5801, debt_to_equity=1.38, equity_multiplier=2.38,
                    interest_coverage=6, pe_ratio=15.33, ps_ratio=1.33, pb_ratio=2.27, book_to_market=0.4405),
    "Utilities": dict(current_ratio=0.93, quick_ratio=0.80, cash_ratio=0.1, gross_margin=0.2840, operating_margin=0.2070,
                       net_margin=0.1174, roa=0.0254, roe=0.0994, asset_turnover=0.22, inventory_turnover=8, dio=46,
                       receivables_turnover=10, dso=37, debt_ratio=0.7445, debt_to_equity=2.91, equity_multiplier=3.91,
                       interest_coverage=3.5, pe_ratio=19.30, ps_ratio=2.44, pb_ratio=2.09, book_to_market=0.4785),
    "Real Estate": dict(current_ratio=2.08, quick_ratio=2.05, cash_ratio=0.2, gross_margin=0.4037, operating_margin=0.2310,
                         net_margin=0.1084, roa=0.0224, roe=0.0550, asset_turnover=0.21, inventory_turnover=None, dio=None,
                         receivables_turnover=15, dso=24, debt_ratio=0.5927, debt_to_equity=1.46, equity_multiplier=2.46,
                         interest_coverage=3, pe_ratio=32.94, ps_ratio=4.41, pb_ratio=2.50, book_to_market=0.4000),
    "Basic Materials": dict(current_ratio=2.37, quick_ratio=1.82, cash_ratio=0.3, gross_margin=0.2780, operating_margin=0.1885,
                             net_margin=0.0969, roa=0.0550, roe=0.1135, asset_turnover=0.57, inventory_turnover=6, dio=61,
                             receivables_turnover=8, dso=46, debt_ratio=0.5154, debt_to_equity=1.06, equity_multiplier=2.06,
                             interest_coverage=6, pe_ratio=22.18, ps_ratio=2.38, pb_ratio=2.85, book_to_market=0.3509),
    "Communication Services": dict(current_ratio=2.40, quick_ratio=2.34, cash_ratio=0.3, gross_margin=0.5182, operating_margin=0.2446,
                                    net_margin=0.2597, roa=0.1260, roe=0.2481, asset_turnover=0.49, inventory_turnover=None,
                                    dio=None, receivables_turnover=8, dso=46, debt_ratio=0.4921, debt_to_equity=0.97,
                                    equity_multiplier=1.97, interest_coverage=5, pe_ratio=29.83, ps_ratio=4.85, pb_ratio=6.82,
                                    book_to_market=0.1466),
    "General / Unknown": dict(current_ratio=1.86, quick_ratio=1.72, cash_ratio=0.3, gross_margin=0.3421, operating_margin=0.1699,
                               net_margin=0.1163, roa=0.0583, roe=0.1476, asset_turnover=0.58, inventory_turnover=7, dio=52,
                               receivables_turnover=9, dso=40, debt_ratio=0.6221, debt_to_equity=2.21, equity_multiplier=3.21,
                               interest_coverage=6, pe_ratio=26.89, ps_ratio=3.15, pb_ratio=4.63, book_to_market=0.2846),
}
# ============================================================
# YFINANCE FETCH LOGIC
# ============================================================
BS_CANDIDATES = {
    "cash": ["Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments", "Cash"],
    "receivables": ["Receivables", "Accounts Receivable", "Net Receivables"],
    "inventory": ["Inventory"],
    "current_assets": ["Current Assets", "Total Current Assets"],
    "total_assets": ["Total Assets"],
    "current_liabilities": ["Current Liabilities", "Total Current Liabilities"],
    "total_liabilities": ["Total Liabilities Net Minority Interest", "Total Liab"],
    "equity": ["Common Stock Equity", "Stockholders Equity", "Total Equity Gross Minority Interest", "Total Stockholder Equity"],
}
IS_CANDIDATES = {
    "revenue": ["Total Revenue", "Operating Revenue"],
    "cogs": ["Cost Of Revenue", "Reconciled Cost Of Revenue"],
    "gross_profit": ["Gross Profit"],
    "ebit": ["Operating Income", "EBIT", "Total Operating Income As Reported"],
    "interest_expense": ["Interest Expense", "Interest Expense Non Operating"],
    "net_income": ["Net Income", "Net Income Common Stockholders", "Net Income Continuous Operations"],
}
SHARES_CANDIDATES = ["Diluted Average Shares", "Basic Average Shares"]
FIELD_LABELS = {
    "cash": "Cash & Cash Equivalents", "receivables": "Accounts Receivable", "inventory": "Inventory",
    "current_assets": "Total Current Assets", "total_assets": "Total Assets",
    "current_liabilities": "Total Current Liabilities", "total_liabilities": "Total Liabilities", "equity": "Shareholders' Equity",
    "revenue": "Net Sales / Revenue", "cogs": "Cost of Goods Sold (COGS)", "gross_profit": "Gross Profit",
    "ebit": "Operating Income (EBIT)", "interest_expense": "Interest Expense", "net_income": "Net Income",
    "share_price": "Share Price (period-end)", "shares_outstanding": "Diluted Shares Outstanding",
}
BS_FIELDS = ["cash", "receivables", "inventory", "current_assets", "total_assets", "current_liabilities",
             "total_liabilities", "equity"]
IS_FIELDS = ["revenue", "cogs", "gross_profit", "ebit", "interest_expense", "net_income"]
def _period_label(col):
    try:
        return f"FY{col.year}"
    except Exception:
        return str(col)
def _lookup(df, candidates, col, missing_list, field_key):
    if df is None or df.empty or col is None:
        missing_list.append(field_key)
        return None
    for cand in candidates:
        if cand in df.index:
            try:
                val = df.loc[cand, col]
            except Exception:
                continue
            if pd.notna(val):
                return float(val)
    missing_list.append(field_key)
    return None
def _nearest_price(hist, target_date, max_days=10):
    """Closest available daily close on or before target_date (falls back
    to the nearest close after it if nothing precedes it)."""
    if hist is None or hist.empty or target_date is None:
        return None
    try:
        target = pd.Timestamp(target_date)
        idx = hist.index
        if getattr(idx, "tz", None) is not None and target.tzinfo is None:
            target = target.tz_localize(idx.tz)
        elif getattr(idx, "tz", None) is None and target.tzinfo is not None:
            target = target.tz_localize(None)
    except Exception:
        return None
    before = hist.index[hist.index <= target]
    if len(before) > 0:
        nearest = before.max()
    else:
        after = hist.index[hist.index >= target]
        if len(after) == 0:
            return None
        nearest = after.min()
    if abs((nearest - target).days) > max_days:
        return None
    try:
        return float(hist.loc[nearest, "Close"])
    except Exception:
        return None
@st.cache_data(ttl=3600, show_spinner=False)
def fetch_company(ticker_symbol, max_periods=MAX_PERIODS):
    """Fetch up to `max_periods` annual periods for one ticker. Never
    raises — returns a dict with an 'error' key set if the fetch failed
    entirely. Each period includes fundamentals + share price + diluted
    shares, so the UI can let the user pick ANY two periods to compare
    (not just the two most recent) without ever needing new code."""
    ticker_symbol = ticker_symbol.strip().upper()
    try:
        t = yf.Ticker(ticker_symbol)
    except Exception as e:
        return dict(ticker=ticker_symbol, error=f"Could not create Ticker object: {e}")
    try:
        info = t.info or {}
    except Exception:
        info = {}
    name = info.get("longName") or info.get("shortName") or ticker_symbol
    sector = info.get("sector") or "Unknown"
    currency = info.get("currency") or "USD"
    fallback_shares = info.get("sharesOutstanding")
    try:
        bs = t.balance_sheet
    except Exception:
        bs = pd.DataFrame()
    try:
        inc = t.income_stmt
    except Exception:
        try:
            inc = t.financials
        except Exception:
            inc = pd.DataFrame()
    if bs is None:
        bs = pd.DataFrame()
    if inc is None:
        inc = pd.DataFrame()
    if bs.empty and inc.empty:
        return dict(ticker=ticker_symbol, error="No financial statement data returned by Yahoo Finance for this ticker.")
    inc_periods_all = sorted(list(inc.columns), reverse=True) if not inc.empty else []
    bs_periods_all = sorted(list(bs.columns), reverse=True) if not bs.empty else []
    master_periods = (inc_periods_all if inc_periods_all else bs_periods_all)[:max_periods]
    def nearest_bs_col(target_date):
        if not bs_periods_all:
            return None
        return min(bs_periods_all, key=lambda d: abs((d - target_date).days))
    try:
        hist = t.history(period="5y", interval="1d", auto_adjust=False)
    except Exception:
        hist = pd.DataFrame()
    periods = []
    for inc_col in master_periods:
        bs_col = nearest_bs_col(inc_col) if bs_periods_all else None
        missing = []
        vals = {}
        for key in BS_FIELDS:
            vals[key] = _lookup(bs, BS_CANDIDATES[key], bs_col, missing, key)
        for key in IS_FIELDS:
            vals[key] = _lookup(inc, IS_CANDIDATES[key], inc_col, missing, key)
        if vals.get("interest_expense") is not None:
            vals["interest_expense"] = abs(vals["interest_expense"])
        shares = _lookup(inc, SHARES_CANDIDATES, inc_col, missing, "shares_outstanding")
        if shares is None and fallback_shares:
            shares = float(fallback_shares)  # approximation: current share count, not historical
        price = _nearest_price(hist, inc_col)
        if price is None:
            missing.append("share_price")
        periods.append(dict(
            label=_period_label(inc_col), date=inc_col, values=vals, missing=missing,
            share_price=price, shares_outstanding=shares,
        ))
    if not periods:
        return dict(ticker=ticker_symbol, error="Could not extract any usable reporting periods for this ticker.")
    return dict(ticker=ticker_symbol, name=name, sector=sector, currency=currency, periods=periods, error=None)
# ============================================================
# EXCEL EXPORT
# ============================================================
def build_excel_report(companies, benchmark_sector, benchmark, matrix_rows, prepared_by=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ratio Comparison"
    title_font = Font(name="Arial", size=14, bold=True)
    subtitle_font = Font(name="Arial", size=10, italic=True, color="666666")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    normal_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    n_cols = 2 + 2 * len(companies) + 1
    last_col_letter = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "Real-Company Ratio Comparison"
    ws["A1"].font = title_font
    ws.merge_cells(f"A2:{last_col_letter}2")
    subtitle_text = f"Companies: {', '.join(c['ticker'] for c in companies)}  |  Benchmark sector: {benchmark_sector}"
    if prepared_by:
        subtitle_text += f"  |  Prepared by: {prepared_by}"
    ws["A2"] = subtitle_text
    ws["A2"].font = subtitle_font
    headers = ["Category", "Ratio"]
    for c in companies:
        headers += [f"{c['ticker']} {c['y1_label']}", f"{c['ticker']} {c['y2_label']}"]
    headers += ["Industry Benchmark"]
    header_row = 4
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    number_formats = {"": "0.00", "%": "0.0%", "x": '0.00"x"', " days": '0" days"'}
    r = header_row
    for row in matrix_rows:
        r += 1
        nf = number_formats.get(row["suffix"], "0.00")
        ws.cell(row=r, column=1, value=row["category"]).font = normal_font
        ws.cell(row=r, column=2, value=row["ratio"]).font = normal_font
        col_idx = 3
        for val in row["company_values"]:
            cell = ws.cell(row=r, column=col_idx, value=("N/A" if val is None else val))
            cell.font = normal_font
            if val is not None:
                cell.number_format = nf
            cell.alignment = Alignment(horizontal="center")
            col_idx += 1
        bcell = ws.cell(row=r, column=col_idx, value=("N/A" if row["benchmark"] is None else row["benchmark"]))
        bcell.font = normal_font
        if row["benchmark"] is not None:
            bcell.number_format = nf
        bcell.alignment = Alignment(horizontal="center")
        for ci in range(1, n_cols + 1):
            ws.cell(row=r, column=ci).border = border
    widths = [16, 28] + [14] * (2 * len(companies)) + [18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C5"
    for c in companies:
        sheet_name = c["ticker"][:31]
        ws2 = wb.create_sheet(sheet_name)
        ws2.merge_cells("A1:C1")
        ws2["A1"] = f"{c['name']} ({c['ticker']}) — Raw Inputs"
        ws2["A1"].font = title_font
        raw_headers = ["Line Item", c["y1_label"], c["y2_label"]]
        for ci, h in enumerate(raw_headers, start=1):
            cell = ws2.cell(row=3, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        rr = 3
        all_keys = BS_FIELDS + IS_FIELDS + ["share_price", "shares_outstanding"]
        for key in all_keys:
            rr += 1
            ws2.cell(row=rr, column=1, value=FIELD_LABELS[key]).font = normal_font
            for ci, yv in ((2, c["raw_y1"].get(key)), (3, c["raw_y2"].get(key))):
                cell = ws2.cell(row=rr, column=ci, value=("N/A" if yv is None else yv))
                cell.font = normal_font
                if yv is not None:
                    cell.number_format = "#,##0.00"
                cell.border = border
            ws2.cell(row=rr, column=1).border = border
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
# ============================================================
# PDF EXPORT
# ============================================================
def build_pdf_report(companies, benchmark_sector, matrix_rows, prepared_by=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(PAGE_SIZE),
        leftMargin=30, rightMargin=30, topMargin=36, bottomMargin=30,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=16, spaceAfter=4)
    sub_style = ParagraphStyle("ReportSub", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=14)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=12, spaceBefore=14, spaceAfter=6,
                                    textColor=colors.HexColor("#1F4E78"))
    footnote_style = ParagraphStyle("Footnote", parent=styles["Normal"], fontSize=7.5, textColor=colors.grey)
    tickers = ", ".join(c["ticker"] for c in companies)
    subtitle_text = f"Companies: {tickers}  |  Benchmark sector: {benchmark_sector}"
    if prepared_by:
        subtitle_text += f"  |  Prepared by: {prepared_by}"
    story = [
        Paragraph("Real-Company Ratio Comparison", title_style),
        Paragraph(subtitle_text, sub_style),
    ]
    col_widths = [140] + [55] * (2 * len(companies)) + [60]
    for cat in CATEGORIES:
        rows = [r for r in matrix_rows if r["category"] == cat]
        if not rows:
            continue
        story.append(Paragraph(cat, section_style))
        header = ["Ratio"]
        for c in companies:
            header += [c["y1_label"], c["y2_label"]]
        header += ["Bench."]
        ticker_header = [""]
        for c in companies:
            ticker_header += [c["ticker"], ""]
        ticker_header += [""]
        data = [ticker_header, header]
        for r in rows:
            line = [r["ratio"]]
            for val in r["company_values"]:
                line.append(fmt(val, r["suffix"]))
            line.append(fmt(r["benchmark"], r["suffix"]))
            data.append(line)
        table = Table(data, colWidths=col_widths[: len(header)], repeatRows=2)
        style_cmds = [
            ("SPAN", (0, 0), (0, 1)),
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 1), colors.white),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        col = 1
        for _ in companies:
            style_cmds.append(("SPAN", (col, 0), (col + 1, 0)))
            col += 2
        table.setStyle(TableStyle(style_cmds))
        story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "Industry benchmark values are general educational reference points, not live data. Market "
        "valuation ratios use the closing share price nearest each period's fiscal year-end. Assessment "
        "thresholds are classroom rules of thumb; real-world norms vary by industry. Generated by the "
        "Real-Company Ratio Comparison Tool.",
        footnote_style,
    ))
    doc.build(story)
    buf.seek(0)
    return buf
# ============================================================
# AI CALC CHALLENGE — ROUND 2 EXPORT (printable submission for grading)
# ============================================================
def _pdf_text(text):
    """Escapes text for safe use inside a reportlab Paragraph (which parses
    a small HTML/XML-like markup, so a literal '<' or '&' pasted from an
    AI response would otherwise break rendering) and converts newlines to
    <br/> so multi-line pasted text still reads as multiple lines."""
    if not text or not text.strip():
        return "(not provided)"
    return xml_escape(text).replace("\n", "<br/>")
def build_calc_challenge_round2_pdf(company, manual_source, ai_prompt, ai_response, rows2, reflection,
                                     prepared_by=None, ai_tool=None, ai_timestamp=None):
    """One-company, submission-ready PDF of the Round 2 (open research)
    four-way comparison — the student's own prediction vs. App (yfinance)
    vs. the student's Manual (10-K) figures vs. the AI's self-sourced
    calculation — plus the supporting evidence (prompt, full AI response,
    which AI tool + when, source citation, reflection) an instructor needs
    to actually grade it rather than just see a table."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(PAGE_SIZE),
        leftMargin=36, rightMargin=36, topMargin=40, bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CCTitle", parent=styles["Title"], fontSize=15, spaceAfter=4)
    sub_style = ParagraphStyle("CCSub", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=12)
    section_style = ParagraphStyle("CCSection", parent=styles["Heading2"], fontSize=12, spaceBefore=14, spaceAfter=6,
                                    textColor=colors.HexColor("#1F4E78"))
    body_style = ParagraphStyle("CCBody", parent=styles["Normal"], fontSize=9, leading=13, spaceAfter=8)
    mono_style = ParagraphStyle("CCMono", parent=styles["Normal"], fontName="Courier", fontSize=7.5, leading=10,
                                 backColor=colors.HexColor("#F2F6FA"), spaceAfter=8, borderPadding=6)
    footnote_style = ParagraphStyle("CCFootnote", parent=styles["Normal"], fontSize=7.5, textColor=colors.grey)

    subtitle = f"{company['name']} ({company['ticker']}) — Round 2: Open Research Challenge"
    if prepared_by:
        subtitle += f"  |  Prepared by: {prepared_by}"
    story = [
        Paragraph("AI Ratio Calculation Challenge", title_style),
        Paragraph(subtitle, sub_style),
    ]

    story.append(Paragraph("Manual (10-K) Source", section_style))
    story.append(Paragraph(_pdf_text(manual_source), body_style))

    story.append(Paragraph("Four-Way Comparison", section_style))
    def _plain_verdict(s):
        # reportlab's default fonts (Helvetica) don't have emoji glyphs — the
        # ✅/❌/⚪ used on-screen would render as blank boxes in the PDF, so
        # swap in plain words for the printed version.
        return s.replace("✅ ", "").replace("❌ ", "").replace("⚪ ", "")
    header = ["Category", "Ratio", "Your\nPrediction", "App\n(yfinance)", "Manual\n(10-K)",
               "AI\n(self-sourced)", "You vs\nApp", "App vs\nManual", "App vs\nAI"]
    data = [header] + [[r["Category"], r["Ratio"], r["Your Prediction"], r["App (yfinance)"],
                        r["Manual (10-K)"], r["AI (self-sourced)"], _plain_verdict(r["You vs App"]),
                        _plain_verdict(r["App vs Manual"]), _plain_verdict(r["App vs AI"])]
                       for r in rows2]
    col_widths = [55, 95, 55, 55, 55, 60, 48, 55, 48]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(table)
    story.append(Spacer(1, 10))

    n_match_you = sum(1 for r in rows2 if r["You vs App"] == "✅ Match")
    n_diff_you = sum(1 for r in rows2 if r["You vs App"] == "❌ Differ")
    n_match_am = sum(1 for r in rows2 if r["App vs Manual"] == "✅ Match")
    n_diff_am = sum(1 for r in rows2 if r["App vs Manual"] == "❌ Differ")
    n_match_aai = sum(1 for r in rows2 if r["App vs AI"] == "✅ Match")
    n_diff_aai = sum(1 for r in rows2 if r["App vs AI"] == "❌ Differ")
    story.append(Paragraph(
        f"You vs App: {n_match_you} match, {n_diff_you} differ &nbsp;|&nbsp; "
        f"App vs Manual: {n_match_am} match, {n_diff_am} differ &nbsp;|&nbsp; "
        f"App vs AI: {n_match_aai} match, {n_diff_aai} differ (5% tolerance)",
        footnote_style,
    ))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Prompt Given to the AI (no data supplied)", section_style))
    story.append(Paragraph(_pdf_text(ai_prompt), mono_style))

    ai_response_heading = "AI's Full Response"
    if ai_tool or ai_timestamp:
        detail = " · ".join(x for x in [ai_tool, (f"recorded {ai_timestamp}" if ai_timestamp else None)] if x)
        ai_response_heading += f" ({detail})"
    story.append(Paragraph(ai_response_heading, section_style))
    story.append(Paragraph(_pdf_text(ai_response), body_style))

    story.append(Paragraph("Student Reflection", section_style))
    story.append(Paragraph(_pdf_text(reflection), body_style))

    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "Generated by the Real-Company Ratio Comparison Tool's AI Calc Challenge (Round 2). "
        "A 5% tolerance is used for the Match/Differ verdicts. This document is a student-submitted "
        "record — the App (yfinance) and Manual (10-K) figures are only as accurate as the data "
        "entered, and the AI response is unverified by the instructor.",
        footnote_style,
    ))
    doc.build(story)
    buf.seek(0)
    return buf
# ============================================================
# AI INTERPRETATION PROMPT (student pastes this into their own AI chat)
# ============================================================
def build_ratio_prompt(companies, benchmark_sector, matrix_rows):
    """Turn the already-computed ratio matrix into a compact text prompt the
    student copies into their own ChatGPT / Gemini / Claude / Perplexity tab.
    Sends COMPUTED RATIOS only (not raw financial statement line items) —
    this app never calls any AI API itself."""
    lines = [
        f"You are a finance teaching assistant. Companies being compared: "
        f"{', '.join(c['ticker'] for c in companies)}. "
        f"Industry benchmark used: {benchmark_sector}.",
        "",
        "Ratio data (Year1 -> Year2, vs benchmark):",
    ]
    for row in matrix_rows:
        vals = []
        ci = 0
        for c in companies:
            v1 = row["company_values"][ci]
            v2 = row["company_values"][ci + 1]
            vals.append(f"{c['ticker']}: {fmt(v1, row['suffix'])} -> {fmt(v2, row['suffix'])}")
            ci += 2
        bench = fmt(row["benchmark"], row["suffix"])
        lines.append(f"- {row['ratio']} ({row['category']}): " + "; ".join(vals) + f" | benchmark: {bench}")
    lines += [
        "",
        "Format your entire response in Markdown, structured exactly like this:",
        "",
        "For EACH company, write a '### TICKER' section header, then under it:",
        "",
        "1. A Markdown TABLE with columns: | Ratio | Latest Value | Benchmark | Assessment |",
        "   Include the 6-8 ratios most relevant to that company's key strengths and "
        "weaknesses, using ONLY the values given above. Never invent numbers. Assessment "
        "column: 'Strength' or 'Weakness' for each row.",
        "",
        "2. Below the table, a '**Notable year-over-year trend:**' line with 1-2 sentences.",
        "",
        "After all companies, add one final '### What to double-check' section with "
        "1-2 sentences on something a student should verify or be skeptical of in this data.",
        "",
        "Keep prose under 300 words total (tables don't count toward that limit).",
        "",
        "If you have web search or browsing available, you may add ONE final short "
        "'### Recent context' section (2-3 sentences per company, optional) noting any "
        "genuinely relevant recent news/events for these tickers. Web search may ONLY be "
        "used for this optional context section — never to change, replace, or supplement "
        "any ratio, benchmark, or numeric value above; those are authoritative as given.",
    ]
    return "\n".join(lines)
# ============================================================
# AI CALC CHALLENGE PROMPT — hands the AI RAW figures (not ratios) and
# asks it to calculate them itself, so its arithmetic can be checked
# against the app's own yfinance-based numbers.
# ============================================================
def build_raw_calc_prompt(company):
    """Builds a prompt for one company's latest comparison year using the
    same post-edit figures (values_y2 / market_cap_y2) the app itself used
    to compute that company's ratios — so any mismatch reflects the AI's
    calculation, not a difference in input data."""
    d = compute_derived(company["values_y2"])
    mkt_cap = company.get("market_cap_y2")
    lines = [
        f"You are being tested on financial ratio calculation for {company['name']} "
        f"({company['ticker']}), fiscal year {company['y2_label']}.",
        "",
        f"Here are the ONLY figures you may use (in {company['currency']}). If a figure "
        "needed for a ratio isn't listed below, say that ratio can't be calculated from "
        "the given data — do NOT invent, estimate, or recall a number from your own "
        "knowledge or training data.",
        "",
    ]
    for key in BS_FIELDS + IS_FIELDS:
        val = d.get(key)
        lines.append(f"- {FIELD_LABELS[key]}: {val:,.0f}" if val is not None else f"- {FIELD_LABELS[key]}: not available")
    lines.append(
        f"- Market Capitalization (share price × diluted shares outstanding): {mkt_cap:,.0f}"
        if mkt_cap is not None else "- Market Capitalization: not available"
    )
    lines += [
        "",
        "For EACH of the following ratios, show your work: the formula, the exact figures "
        "from the list above you plugged in, and the calculated value.",
        "",
    ]
    for rdef in RATIO_DEFS:
        lines.append(f"- {rdef['name']} ({rdef['category']}): {rdef['formula']}")
    lines += [
        "",
        "After showing your work for all of them, finish with a section titled "
        "'### Final Answers' containing a single Markdown table with exactly these "
        "columns: | Ratio Key | Calculated Value |",
        "Use exactly these ratio keys, one per row, in this order: "
        + ", ".join(r["key"] for r in RATIO_DEFS) + ".",
        "",
        "For the Calculated Value column, give just the plain number (e.g. 1.86, 24.5, "
        "45, 3.87) with no percent sign, 'x' suffix, or 'days' text — the student will "
        "apply the correct units when transcribing your answer. If a ratio can't be "
        "calculated from the given data, write N/A for that row.",
    ]
    return "\n".join(lines)
# ============================================================
# AI OPEN RESEARCH PROMPT (Round 2) — no figures given at all. The AI has
# to find the company's own latest financials itself AND calculate the
# ratios, and must state exactly what period/source it used so a mismatch
# can be diagnosed as a data problem, a math problem, or both.
# ============================================================
def build_open_research_prompt(company):
    lines = [
        f"You are being tested on financial research AND ratio calculation for "
        f"{company['name']} ({company['ticker']}).",
        "",
        "Do NOT ask me for any figures — find them yourself, using whatever "
        "training knowledge, web search, or browsing tools you have available. "
        "Use the company's most recently completed full fiscal year (not a "
        "partial or trailing-twelve-month period, unless that's genuinely the "
        "most recent full year available to you).",
        "",
        "Before anything else, write a '### Data Source & Period' section stating: "
        "(1) the exact fiscal year and fiscal year-end date of the figures you're "
        "using, and (2) where those figures came from (e.g. the company's 10-K, a "
        "specific database, your training knowledge with an approximate as-of "
        "date). Be specific — 'a financial database' on its own is not specific "
        "enough.",
        "",
        "Then, for EACH of the following ratios, show your work: the formula, the "
        "exact raw figures you found and are using (state them explicitly, with "
        "their fiscal period), and the calculated value.",
        "",
    ]
    for rdef in RATIO_DEFS:
        lines.append(f"- {rdef['name']} ({rdef['category']}): {rdef['formula']}")
    lines += [
        "",
        "After showing your work for all of them, finish with a section titled "
        "'### Final Answers' containing a single Markdown table with exactly these "
        "columns: | Ratio Key | Calculated Value |",
        "Use exactly these ratio keys, one per row, in this order: "
        + ", ".join(r["key"] for r in RATIO_DEFS) + ".",
        "",
        "For the Calculated Value column, give just the plain number (e.g. 1.86, "
        "24.5, 45, 3.87) with no percent sign, 'x' suffix, or 'days' text. If you "
        "can't find reliable data for a ratio, write N/A rather than guessing.",
    ]
    return "\n".join(lines)
# ============================================================
# STUDENT LOGIN
# ============================================================
def load_credentials():
    """Loads the class roster used to gate access to this app.

    Two sources are supported, checked in this order:
      1. Streamlit secrets (st.secrets) — used for Streamlit Community
         Cloud deployments. Configure a [credentials.usernames.<user>]
         table and a [cookie] table in the app's Secrets settings.
      2. A local config.yaml file next to this script — used when
         running the app on your own machine. See build_roster.py
         and README.md for how to generate it from a plain roster
         CSV (never hand-edit real password hashes).

    Both formats mirror streamlit-authenticator's expected structure:
        credentials:
          usernames:
            jdoe:
              name: Jane Doe
              email: jdoe@school.edu
              password: <bcrypt hash>
        cookie:
          name: ratio_tool_auth
          key: <a long random secret string>
          expiry_days: 7
    """
    try:
        if "credentials" in st.secrets:
            creds = {
                "usernames": {
                    u: dict(v) for u, v in st.secrets["credentials"]["usernames"].items()
                }
            }
            cookie_cfg = st.secrets.get("cookie", {})
            return (
                creds,
                cookie_cfg.get("name", "ratio_tool_auth"),
                cookie_cfg.get("key", "please-change-this-cookie-secret"),
                float(cookie_cfg.get("expiry_days", 7)),
            )
    except Exception:
        pass  # st.secrets not configured — fall through to local config.yaml
    if os.path.exists("config.yaml"):
        with open("config.yaml") as cf:
            config = yaml.load(cf, Loader=stauth.SafeLoader)
        cookie_cfg = config.get("cookie", {})
        return (
            config["credentials"],
            cookie_cfg.get("name", "ratio_tool_auth"),
            cookie_cfg.get("key", "please-change-this-cookie-secret"),
            float(cookie_cfg.get("expiry_days", 7)),
        )
    st.error(
        "No login roster found. The instructor needs to create **config.yaml** "
        "(run `python build_roster.py` on a roster.csv) or configure Secrets "
        "on Streamlit Community Cloud. See README.md for step-by-step instructions."
    )
    st.stop()

st.title("🏢 Real-Company Ratio Comparison Tool")
st.caption("Pull real financial statements, compare companies over the years you choose, and benchmark against industry norms.")

_credentials, _cookie_name, _cookie_key, _cookie_expiry = load_credentials()
authenticator = stauth.Authenticate(_credentials, _cookie_name, _cookie_key, _cookie_expiry)
authenticator.login(location="main")

_auth_status = st.session_state.get("authentication_status")
if _auth_status is False:
    st.error("Username or password is incorrect. Contact your instructor if you've forgotten your credentials.")
    st.stop()
elif _auth_status is None:
    st.info("Please log in with the username and password your instructor gave you to use this tool.")
    st.stop()

# Authenticated from this point on.
student_name = st.session_state.get("name") or st.session_state.get("username")
student_username = st.session_state.get("username")

# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.success(f"Logged in as **{student_name}**")
    authenticator.logout("Log out", "sidebar")
    st.divider()
    team_members_input = st.text_input(
        "Team members for this submission (comma-separated, up to 4)",
        value=load_team_members(student_username) or student_name,
        key="team_members_input",
        help="This is what appears as 'Prepared by' on every downloaded "
             "report (Excel/PDF) — defaults to just you, add teammates' "
             "names here if this is group work. Remembered for next time "
             "you log in.",
    )
    if st.session_state.get("_team_members_saved_value") != team_members_input:
        save_team_members(student_username, team_members_input)
        st.session_state["_team_members_saved_value"] = team_members_input
    prepared_by_display, _team_truncated = compute_prepared_by(team_members_input, student_name)
    if _team_truncated:
        st.caption(f"⚠️ Only the first {MAX_TEAM_MEMBERS} names will appear on downloaded reports.")
    st.divider()
    st.header("📘 How to use this tool")
    st.markdown(
        """
        1. Enter up to **4 stock tickers** (comma-separated) and click
           **Fetch Data**. Figures come from Yahoo Finance via the
           free `yfinance` library — up to 4 recent annual periods
           per company are pulled.
        2. For each company, **pick which two years to compare** from
           the dropdowns (defaults to the two most recent). This is
           why the same app keeps working next year: just fetch again
           and pick the newer years — no code changes ever needed.
        3. **Always review the editable tables** — free data sources
           sometimes miss a line item or label it unexpectedly. Fix
           anything that looks wrong before trusting the ratios.
        4. Pick (or keep the auto-detected) **industry benchmark**
           and adjust its values if you have more precise figures.
        5. Browse the ratio tabs — including **Market Valuation**
           (P/E, P/S, P/B, Book-to-Market) — then download the
           Summary Dashboard as CSV, Excel, or PDF.
        *Data quality note:* yfinance is free and convenient but not
        always complete. Cross-check anything surprising against the
        company's actual 10-K/annual report.
        """
    )
    st.caption("Ratio thresholds and benchmark values are illustrative educational references, not official current industry data.")
    st.divider()
    st.markdown("**AI Interpretation**")
    st.caption(
        "This app never calls an AI API itself and never asks for an API key. "
        "In the Summary Dashboard tab, you'll get a ready-made prompt to copy "
        "into your own ChatGPT, Gemini, Claude, or Perplexity tab — using "
        "whichever one you're already logged into — then paste the answer "
        "back in to compare it against the numbers above."
    )
    st.divider()
    with st.expander("🛠️ About this tool"):
        st.caption(
            "This tool was designed and developed individually by the instructor. AI "
            "assistance (Claude Sonnet 5, Anthropic) was used selectively — not as a "
            "substitute for original development, but to pressure-test the finished "
            "code against edge cases (concurrent logins, malformed pasted text, PDF "
            "rendering failures), review implementation choices (e.g., Streamlit "
            "session-state patterns, login/auth libraries), and suggest refinements. "
            "All AI-suggested changes were critically evaluated, and every line of "
            "functionality was reviewed, tested, and verified by the instructor before "
            "use in class. No student data — including names, submissions, grades, or "
            "any other personally identifiable information — was shared with the AI "
            "at any point during development or testing."
        )
# ============================================================
# TICKER INPUT & FETCH
# ============================================================
default_tickers = "AAPL, MSFT"
tickers_input = st.text_input("Stock tickers (comma-separated, up to 4)", value=default_tickers)
fetch_clicked = st.button("🔄 Fetch Data", type="primary")
tickers = [t.strip().upper() for t in tickers_input.split(",") if t.strip()][:4]
if fetch_clicked or "fetched_companies" not in st.session_state:
    if tickers:
        with st.spinner("Fetching financial statements from Yahoo Finance..."):
            fetched = [fetch_company(t) for t in tickers]
        st.session_state["fetched_companies"] = fetched
        st.session_state["fetched_tickers"] = tickers
companies_raw = st.session_state.get("fetched_companies", [])
errors = [c for c in companies_raw if c.get("error")]
companies_raw = [c for c in companies_raw if not c.get("error")]
for e in errors:
    st.error(f"**{e['ticker']}**: {e['error']}")
if not companies_raw:
    st.info("Enter one or more tickers above and click **Fetch Data** to get started.")
    st.stop()
# ============================================================
# REVIEW & EDIT FETCHED DATA
# ============================================================
st.subheader("📝 Review & Edit Fetched Data")
st.write(
    "Pick which two years to compare for each company, then check the auto-fetched figures against "
    "the company's real filings. Any field yfinance couldn't find is flagged so you know to fill it "
    "in by hand."
)
companies = []
for c in companies_raw:
    period_labels = [p["label"] for p in c["periods"]]  # most-recent first
    period_by_label = {p["label"]: p for p in c["periods"]}
    with st.expander(f"{c['name']} ({c['ticker']}) — Sector: {c['sector']}", expanded=(len(companies_raw) <= 2)):
        yr1, yr2 = st.columns(2)
        default_y2_idx = 0
        default_y1_idx = 1 if len(period_labels) > 1 else 0
        with yr1:
            y1_label = st.selectbox(f"Year 1 (baseline) — {c['ticker']}", period_labels,
                                     index=default_y1_idx, key=f"y1_sel_{c['ticker']}")
        with yr2:
            y2_label = st.selectbox(f"Year 2 (comparison) — {c['ticker']}", period_labels,
                                     index=default_y2_idx, key=f"y2_sel_{c['ticker']}")
        p1, p2 = period_by_label[y1_label], period_by_label[y2_label]
        col_bs, col_is = st.columns(2)
        with col_bs:
            st.markdown("**Balance Sheet**")
            bs_df = pd.DataFrame({
                "Item": [FIELD_LABELS[k] for k in BS_FIELDS],
                y1_label: [p1["values"].get(k) if p1["values"].get(k) is not None else 0.0 for k in BS_FIELDS],
                y2_label: [p2["values"].get(k) if p2["values"].get(k) is not None else 0.0 for k in BS_FIELDS],
            })
            edited_bs = st.data_editor(
                bs_df, key=f"bs_editor_{c['ticker']}_{y1_label}_{y2_label}", hide_index=True,
                num_rows="fixed", use_container_width=True,
                column_config={
                    "Item": st.column_config.TextColumn("Line Item", disabled=True),
                    y1_label: st.column_config.NumberColumn(y1_label, format="%.0f"),
                    y2_label: st.column_config.NumberColumn(y2_label, format="%.0f"),
                },
            )
        with col_is:
            st.markdown("**Income Statement**")
            is_df = pd.DataFrame({
                "Item": [FIELD_LABELS[k] for k in IS_FIELDS],
                y1_label: [p1["values"].get(k) if p1["values"].get(k) is not None else 0.0 for k in IS_FIELDS],
                y2_label: [p2["values"].get(k) if p2["values"].get(k) is not None else 0.0 for k in IS_FIELDS],
            })
            edited_is = st.data_editor(
                is_df, key=f"is_editor_{c['ticker']}_{y1_label}_{y2_label}", hide_index=True,
                num_rows="fixed", use_container_width=True,
                column_config={
                    "Item": st.column_config.TextColumn("Line Item", disabled=True),
                    y1_label: st.column_config.NumberColumn(y1_label, format="%.0f"),
                    y2_label: st.column_config.NumberColumn(y2_label, format="%.0f"),
                },
            )
        st.markdown("**Market Data** (used for P/E, P/S, P/B, Book-to-Market)")
        st.caption(
            "Share price is the closing price nearest each period's fiscal year-end. If diluted shares "
            "weren't found for a historical period, the company's current share count is used as an "
            "approximation — edit it if you know the actual figure for that year."
        )
        market_df = pd.DataFrame({
            "Item": ["Share Price (period-end)", "Diluted Shares Outstanding"],
            y1_label: [p1["share_price"] or 0.0, p1["shares_outstanding"] or 0.0],
            y2_label: [p2["share_price"] or 0.0, p2["shares_outstanding"] or 0.0],
        })
        edited_mkt = st.data_editor(
            market_df, key=f"mkt_editor_{c['ticker']}_{y1_label}_{y2_label}", hide_index=True,
            num_rows="fixed", use_container_width=True,
            column_config={
                "Item": st.column_config.TextColumn("Item", disabled=True),
                y1_label: st.column_config.NumberColumn(y1_label, format="%.2f"),
                y2_label: st.column_config.NumberColumn(y2_label, format="%.2f"),
            },
        )
        combined_missing = sorted(set(p1["missing"]) | set(p2["missing"]))
        if combined_missing:
            st.warning("⚠️ Not found automatically for one or both years (please review): " +
                       ", ".join(FIELD_LABELS[k] for k in combined_missing))
        else:
            st.success("✅ All fields found automatically.")
        raw_y1 = {k: p1["values"].get(k) for k in BS_FIELDS + IS_FIELDS}
        raw_y1["share_price"] = p1["share_price"]
        raw_y1["shares_outstanding"] = p1["shares_outstanding"]
        raw_y2 = {k: p2["values"].get(k) for k in BS_FIELDS + IS_FIELDS}
        raw_y2["share_price"] = p2["share_price"]
        raw_y2["shares_outstanding"] = p2["shares_outstanding"]
        values_y1 = {k: float(v) for k, v in zip(BS_FIELDS, edited_bs[y1_label])}
        values_y1.update({k: float(v) for k, v in zip(IS_FIELDS, edited_is[y1_label])})
        values_y2 = {k: float(v) for k, v in zip(BS_FIELDS, edited_bs[y2_label])}
        values_y2.update({k: float(v) for k, v in zip(IS_FIELDS, edited_is[y2_label])})
        price1, shares1 = float(edited_mkt[y1_label][0]), float(edited_mkt[y1_label][1])
        price2, shares2 = float(edited_mkt[y2_label][0]), float(edited_mkt[y2_label][1])
        market_cap_y1 = (price1 * shares1) if (price1 > 0 and shares1 > 0) else None
        market_cap_y2 = (price2 * shares2) if (price2 > 0 and shares2 > 0) else None
        companies.append(dict(
            ticker=c["ticker"], name=c["name"], sector=c["sector"], currency=c["currency"],
            y1_label=y1_label, y2_label=y2_label,
            values_y1=values_y1, values_y2=values_y2,
            market_cap_y1=market_cap_y1, market_cap_y2=market_cap_y2,
            raw_y1=raw_y1, raw_y2=raw_y2,
        ))
for c in companies:
    d1 = compute_derived(c["values_y1"])
    d2 = compute_derived(c["values_y2"])
    c["ratios_y1"] = compute_ratios(d1, market_cap=c["market_cap_y1"])
    c["ratios_y2"] = compute_ratios(d2, market_cap=c["market_cap_y2"])
# ============================================================
# INDUSTRY BENCHMARK SELECTION
# ============================================================
st.subheader("📊 Industry Benchmark")
sector_options = list(BENCHMARKS.keys())
detected_sector = companies[0]["sector"] if companies[0]["sector"] in BENCHMARKS else "General / Unknown"
default_idx = sector_options.index(detected_sector) if detected_sector in sector_options else len(sector_options) - 1
benchmark_sector = st.selectbox(
    f"Benchmark sector (auto-detected from {companies[0]['ticker']}: {companies[0]['sector']})",
    sector_options, index=default_idx,
)
st.caption(
    f"📊 = real sector data (Finviz sector aggregates, as of {BENCHMARK_AS_OF}) or a DuPont-derived "
    "value calculated from that real data. 📝 = illustrative estimate (Finviz doesn't publish this "
    "ratio at the sector level). Edit any value below if you have more precise figures (e.g. from "
    "Damodaran Online at NYU Stern, CSIMarket, or IBISWorld)."
)
bench_defaults = BENCHMARKS[benchmark_sector]
bench_df = pd.DataFrame({
    "Ratio": [r["name"] for r in RATIO_DEFS],
    "Source": ["📊 Real (Finviz)" if r["key"] in REAL_RATIO_KEYS else "📝 Illustrative" for r in RATIO_DEFS],
    "Format": [{"": "plain", "%": "percent (enter as fraction, e.g. 0.35)", "x": "multiple",
                " days": "days"}[r["suffix"]] for r in RATIO_DEFS],
    "Benchmark Value": [bench_defaults.get(r["key"]) for r in RATIO_DEFS],
})
edited_bench = st.data_editor(
    bench_df, key=f"bench_editor_{benchmark_sector}", hide_index=True, num_rows="fixed", use_container_width=True,
    column_config={
        "Ratio": st.column_config.TextColumn(disabled=True),
        "Source": st.column_config.TextColumn(disabled=True),
        "Format": st.column_config.TextColumn(disabled=True),
        "Benchmark Value": st.column_config.NumberColumn(format="%.4f"),
    },
)
benchmark = {r["key"]: (float(v) if pd.notna(v) else None) for r, v in zip(RATIO_DEFS, edited_bench["Benchmark Value"])}
# ============================================================
# RENDERING HELPERS
# ============================================================
def render_ratio_comparison(rdef):
    st.markdown(f"#### {rdef['name']}")
    st.caption(f"Formula: {rdef['formula']}")
    st.caption(rdef["guide"])
    src = ("📊 Benchmark is real sector data (Finviz)" if rdef["key"] in REAL_RATIO_KEYS
           else "📝 Benchmark is an illustrative estimate")
    st.caption(src)
    is_valuation = rdef["category"] == "Market Valuation"
    labels = LEVEL_LABEL_VALUATION if is_valuation else LEVEL_LABEL
    bench_val = benchmark.get(rdef["key"])
    rows = []
    chart_data = {}
    for c in companies:
        v1 = c["ratios_y1"][rdef["key"]]
        v2 = c["ratios_y2"][rdef["key"]]
        delta = (v2 - v1) if (v1 is not None and v2 is not None) else None
        vs_bench = (v2 - bench_val) if (v2 is not None and bench_val is not None) else None
        level = rdef["level_fn"](v2)
        rows.append({
            "Company": f"{c['ticker']}", c["y1_label"]: fmt(v1, rdef["suffix"]), c["y2_label"]: fmt(v2, rdef["suffix"]),
            "YoY Change": fmt_delta(delta, rdef["suffix"]) or "N/A",
            "vs Benchmark": fmt_delta(vs_bench, rdef["suffix"]) or "N/A",
            "Assessment": labels[level],
        })
        if v2 is not None:
            chart_data[c["ticker"]] = v2
    if bench_val is not None:
        chart_data["Benchmark"] = bench_val
    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)
    if chart_data:
        chart_df = pd.DataFrame({"Latest Value": chart_data})
        st.bar_chart(chart_df)
    st.write("")
def render_category_tab(category):
    defs = [r for r in RATIO_DEFS if r["category"] == category]
    for rdef in defs:
        render_ratio_comparison(rdef)
        st.divider()
# ============================================================
# TABS
# ============================================================
tab_liq, tab_prof, tab_eff, tab_solv, tab_market, tab_summary, tab_ai_calc = st.tabs(
    ["💧 Liquidity", "💰 Profitability", "⚙️ Efficiency", "🏦 Solvency", "📈 Market Valuation",
     "📋 Summary Dashboard", "🧪 AI Calc Challenge"]
)
with tab_liq:
    st.write("Liquidity ratios measure whether a company can meet its short-term obligations.")
    render_category_tab("Liquidity")
with tab_prof:
    st.write("Profitability ratios measure how well a company converts sales and assets into profit.")
    render_category_tab("Profitability")
with tab_eff:
    st.write("Efficiency ratios measure how well a company uses its assets to generate sales.")
    render_category_tab("Efficiency")
with tab_solv:
    st.write("Solvency ratios measure long-term financial risk and reliance on debt financing.")
    render_category_tab("Solvency")
with tab_market:
    st.write(
        "Market valuation ratios relate a company's stock price to its fundamentals. **Unlike the "
        "ratios in the other tabs, low/high here doesn't mean bad/good** — it reflects how the market "
        "is pricing the stock relative to earnings, sales, or book value, which depends heavily on "
        "growth expectations and perceived risk, not just fundamental quality."
    )
    st.caption("Prices used are the closing price nearest each period's fiscal year-end, not necessarily today's price.")
    render_category_tab("Market Valuation")
with tab_summary:
    st.subheader("📋 Summary Dashboard")
    st.caption(f"Companies: {', '.join(c['ticker'] for c in companies)}  |  Benchmark: {benchmark_sector}")
    matrix_rows = []
    for rdef in RATIO_DEFS:
        company_values = []
        for c in companies:
            company_values.append(c["ratios_y1"][rdef["key"]])
            company_values.append(c["ratios_y2"][rdef["key"]])
        matrix_rows.append(dict(
            category=rdef["category"], ratio=rdef["name"], suffix=rdef["suffix"],
            company_values=company_values, benchmark=benchmark.get(rdef["key"]),
        ))
    display_cols = {"Category": [], "Ratio": []}
    for c in companies:
        display_cols[f"{c['ticker']} {c['y1_label']}"] = []
        display_cols[f"{c['ticker']} {c['y2_label']}"] = []
    display_cols["Industry Benchmark"] = []
    for row in matrix_rows:
        display_cols["Category"].append(row["category"])
        display_cols["Ratio"].append(row["ratio"])
        ci = 0
        for c in companies:
            display_cols[f"{c['ticker']} {c['y1_label']}"].append(fmt(row["company_values"][ci], row["suffix"]))
            display_cols[f"{c['ticker']} {c['y2_label']}"].append(fmt(row["company_values"][ci + 1], row["suffix"]))
            ci += 2
        display_cols["Industry Benchmark"].append(fmt(row["benchmark"], row["suffix"]))
    df_display = pd.DataFrame(display_cols)
    st.dataframe(df_display, use_container_width=True, hide_index=True)
    st.markdown("#### ⬇️ Download this comparison")
    st.caption(f"Prepared by: {prepared_by_display}")
    dl1, dl2, dl3 = st.columns(3)
    with dl1:
        csv_bytes = df_display.to_csv(index=False).encode("utf-8")
        st.download_button("CSV", data=csv_bytes, file_name="company_ratio_comparison.csv",
                            mime="text/csv", use_container_width=True)
    with dl2:
        excel_buf = build_excel_report(companies, benchmark_sector, benchmark, matrix_rows, prepared_by=prepared_by_display)
        st.download_button("Excel (.xlsx)", data=excel_buf, file_name="company_ratio_comparison.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True)
    with dl3:
        pdf_buf = build_pdf_report(companies, benchmark_sector, matrix_rows, prepared_by=prepared_by_display)
        st.download_button("PDF Report", data=pdf_buf, file_name="company_ratio_comparison.pdf",
                            mime="application/pdf", use_container_width=True)
    st.divider()
    st.subheader("🤖 AI Interpretation — bring your own chat")
    st.caption(
        "This app doesn't call any AI itself. Instead: copy the prompt below, open "
        "your own AI chat in a new tab (already logged into whichever one you use), "
        "paste the prompt in, then paste its answer back here to compare against "
        "the numbers above. Treat the AI's answer the same way you'd treat a "
        "classmate's first draft: a useful starting point, not a verified answer."
    )
    current_signature = (
        tuple(c["ticker"] for c in companies),
        tuple((c["y1_label"], c["y2_label"]) for c in companies),
        benchmark_sector,
    )
    prompt_text = build_ratio_prompt(companies, benchmark_sector, matrix_rows)

    st.markdown("**Step 1 — open your AI chat of choice** (opens in a new tab, using your own login):")
    chat_links = [
        ("ChatGPT", "https://chatgpt.com/"),
        ("Gemini", "https://gemini.google.com/"),
        ("Claude", "https://claude.ai/new"),
        ("Perplexity", "https://www.perplexity.ai/"),
    ]
    link_cols = st.columns(len(chat_links))
    for col, (label, url) in zip(link_cols, chat_links):
        with col:
            st.link_button(label, url, use_container_width=True)

    st.markdown("**Step 2 — copy this prompt** (hover the box, click the copy icon) and paste it into the chat you opened:")
    st.code(prompt_text, language="text")

    st.markdown("**Step 3 — paste the AI's answer back here:**")
    pasted_response = st.text_area(
        "Paste the full response from ChatGPT / Gemini / Claude / Perplexity here",
        key="pasted_ai_response", height=220,
        placeholder="Paste the AI's answer here...",
    )
    if pasted_response.strip():
        st.session_state["ai_analysis_text"] = pasted_response
        st.session_state["ai_prompt_text"] = prompt_text
        st.session_state["ai_analysis_signature"] = current_signature

    if st.session_state.get("ai_analysis_text"):
        if st.session_state.get("ai_analysis_signature") != current_signature:
            st.warning(
                "⚠️ The company/year/benchmark selection has changed since this "
                "response was pasted in — regenerate the prompt (Step 2) and get "
                "a fresh answer so it matches what's currently shown above."
            )
        st.markdown("**Pasted AI response, rendered:**")
        with st.container(border=True):
            st.markdown(st.session_state["ai_analysis_text"])
        with st.expander("📋 Prompt this response was generated from (for your disclosure statement)"):
            st.code(st.session_state["ai_prompt_text"], language="text")
        st.markdown("**Your verification** — required before you use this in an assignment:")
        st.text_area(
            "Does this match your own read of the numbers? What, if anything, "
            "would you correct or add?",
            key="student_ai_critique", height=150,
        )
    st.caption(
        "Reminder: industry benchmark values are illustrative educational reference points, not live "
        "data. Market valuation ratios use period-end share prices. Always cross-check unusual figures "
        "against the company's actual financial filings."
    )
# ============================================================
# AI CALC CHALLENGE TAB
# ============================================================
with tab_ai_calc:
    st.subheader("🧪 AI Ratio Calculation Challenge")
    st.caption(
        "This is a different exercise from the AI Interpretation section in the Summary "
        "Dashboard tab. There, the AI reads ratios the app already calculated and comments "
        "on them. Here, across two rounds, you test whether the AI can calculate — and even "
        "research — the ratios itself. AI models are generally good at explaining ratios but "
        "do sometimes get the data or the math wrong, especially on trickier formulas like "
        "interest coverage, DSO, or book-to-market — that's exactly what this is meant to catch."
    )
    if not progress_store_available():
        st.caption(
            "ℹ️ Autosave isn't enabled for this class yet, so your answers below only last for "
            "this browser tab — download your CSV/PDF before you close it so you don't lose your "
            "work. (Instructor: see README.md → 'Persistent progress' to turn autosave on.)"
        )
    for c in companies:
        with st.expander(f"{c['name']} ({c['ticker']}) — {c['y2_label']}", expanded=(len(companies) <= 2)):
            ticker, y2 = c["ticker"], c["y2_label"]
            saved = load_progress_for_company(student_username, ticker, y2)
            st.markdown("### Round 1 — AI calculates from data you give it")
            st.caption(
                "The AI gets the exact same figures the app used for this company — so any "
                "mismatch below can only mean one thing: the AI's arithmetic or formula was wrong."
            )
            calc_prompt = build_raw_calc_prompt(c)
            st.markdown(
                "**Step 1 — copy this prompt** and paste it into your own AI chat "
                "(use the links in the Summary Dashboard tab, or reuse a tab you already have open):"
            )
            st.code(calc_prompt, language="text")
            st.markdown("**Step 2 — paste the AI's full response here** (for your disclosure record):")
            ai_calc_response_val = st.text_area(
                "Full AI response",
                value=saved.get("ai_calc_response") or "",
                key=f"ai_calc_response_{c['ticker']}", height=200,
                placeholder="Paste the AI's full response here...",
            )
            ai_calc_tool, ai_calc_timestamp = record_ai_disclosure(
                "r1", c["ticker"], ai_calc_response_val,
                saved_tool=saved.get("r1_tool"), saved_timestamp=saved.get("r1_stamp"),
            )
            st.markdown(
                "**Step 3 — transcribe the AI's 'Final Answers' table below**, matching each "
                "ratio to the plain number it gave you (leave blank if it answered N/A):"
            )
            saved_calc_values = saved.get("ai_calc_values", {})
            calc_df = pd.DataFrame({
                "Ratio": [r["name"] for r in RATIO_DEFS],
                "Category": [r["category"] for r in RATIO_DEFS],
                "AI's Value": [saved_calc_values.get(r["key"]) for r in RATIO_DEFS],
            })
            edited_calc = st.data_editor(
                calc_df, key=f"ai_calc_editor_{c['ticker']}_{c['y2_label']}", hide_index=True,
                num_rows="fixed", use_container_width=True,
                column_config={
                    "Ratio": st.column_config.TextColumn(disabled=True),
                    "Category": st.column_config.TextColumn(disabled=True),
                    "AI's Value": st.column_config.NumberColumn(
                        "AI's Value (plain number — no %, x, or 'days')", format="%.4f"
                    ),
                },
            )
            rows = []
            for rdef, ai_val in zip(RATIO_DEFS, edited_calc["AI's Value"]):
                app_raw = c["ratios_y2"][rdef["key"]]
                app_disp = display_magnitude(app_raw, rdef["suffix"])
                ai_disp = float(ai_val) if pd.notna(ai_val) else None
                verdict = compare_flag(app_disp, ai_disp)
                rows.append({
                    "Category": rdef["category"],
                    "Ratio": rdef["name"],
                    "App (yfinance) Value": fmt(app_raw, rdef["suffix"]),
                    "AI's Value": (f"{ai_disp:,.2f}" if ai_disp is not None else "N/A"),
                    "Result": verdict,
                })
            cmp_df = pd.DataFrame(rows)
            st.markdown("**Comparison — App (yfinance-based) vs. AI's calculation:**")
            st.dataframe(cmp_df, use_container_width=True, hide_index=True)
            n_match = sum(1 for r in rows if r["Result"] == "✅ Match")
            n_mismatch = sum(1 for r in rows if r["Result"] == "❌ Differ")
            n_na = sum(1 for r in rows if r["Result"] == "⚪ N/A")
            st.caption(f"✅ {n_match} matched · ❌ {n_mismatch} mismatched · ⚪ {n_na} not answered/entered yet")
            st.markdown("**Your reflection on Round 1** — required:")
            st.text_area(
                "Where (if anywhere) did the AI's numbers differ from the app's, and why do "
                "you think that happened? (e.g. wrong line item, arithmetic slip, unit confusion)",
                value=saved.get("ai_calc_reflection") or "",
                key=f"ai_calc_reflection_{c['ticker']}", height=120,
            )

            st.divider()
            st.markdown("### Round 2 — Open research: you and the AI both go find the real numbers")
            st.caption(
                "This time nobody gets any figures handed to them. You go find this company's "
                "actual financial statements yourself (its 10-K or annual report) and calculate "
                "the ratios by hand. Separately, the AI has to research the SAME company on its "
                "own — no data given — and calculate the ratios itself, stating exactly which "
                "period and source it used. You'll end up with three independent answers for "
                "each ratio: the app's (yfinance), yours (10-K, by hand), and the AI's "
                "(self-researched) — which is the real test of whether any of them can be trusted "
                "blindly."
            )
            st.markdown(
                "**Step 0 — before you look anything up**, write your own gut-check guess for as "
                "many ratios as you're willing to estimate (a partial guess is fine — leave the "
                "rest blank):"
            )
            saved_predictions = saved.get("predictions", {})
            predict_df = pd.DataFrame({
                "Ratio": [r["name"] for r in RATIO_DEFS],
                "Category": [r["category"] for r in RATIO_DEFS],
                "Your Prediction": [saved_predictions.get(r["key"]) for r in RATIO_DEFS],
            })
            edited_predict = st.data_editor(
                predict_df, key=f"predict_editor_{c['ticker']}_{c['y2_label']}", hide_index=True,
                num_rows="fixed", use_container_width=True,
                column_config={
                    "Ratio": st.column_config.TextColumn(disabled=True),
                    "Category": st.column_config.TextColumn(disabled=True),
                    "Your Prediction": st.column_config.NumberColumn(
                        "Your Prediction (plain number — no %, x, or 'days')", format="%.4f"
                    ),
                },
            )
            predictions = dict(zip([r["key"] for r in RATIO_DEFS], edited_predict["Your Prediction"]))

            predict_lock_key = f"predict_confirmed_{c['ticker']}"

            # Autosave Round 1 + predictions now (Round 2's own fields, if any
            # were saved previously, are carried forward via **saved so this
            # doesn't wipe them out while the prediction gate is still locked).
            autosave_progress(student_username, ticker, y2, {
                **saved,
                "ai_calc_response": ai_calc_response_val,
                "r1_tool": ai_calc_tool, "r1_stamp": ai_calc_timestamp,
                "ai_calc_values": {r["key"]: (float(v) if pd.notna(v) else None)
                                    for r, v in zip(RATIO_DEFS, edited_calc["AI's Value"])},
                "ai_calc_reflection": st.session_state.get(f"ai_calc_reflection_{c['ticker']}", ""),
                "predictions": {k: (float(v) if pd.notna(v) else None) for k, v in predictions.items()},
                "predict_confirmed": st.session_state.get(predict_lock_key, False),
            })

            if not st.session_state.get(predict_lock_key, False):
                st.info(
                    "🔒 The 10-K lookup, AI prompt, and comparison below stay locked until you "
                    "confirm your predictions — this is meant to be your own gut-check, made "
                    "before you or the AI look anything up, not a value you fill in afterward to "
                    "match what you see below."
                )
                if st.button(
                    "I've made my predictions — unlock the rest of Round 2",
                    key=f"predict_unlock_{c['ticker']}",
                ):
                    st.session_state[predict_lock_key] = True
                    st.rerun()
                continue
            st.markdown(
                "**Step 1 — find the real numbers.** Look up "
                f"{c['name']}'s ({c['ticker']}) most recent 10-K or annual report (e.g. via "
                "[sec.gov EDGAR](https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany) or "
                "the company's investor relations site) and enter the figures you find below:"
            )
            manual_keys = BS_FIELDS + IS_FIELDS + ["share_price", "shares_outstanding"]
            saved_manual = saved.get("manual_values", {})
            manual_df = pd.DataFrame({
                "Line Item": [FIELD_LABELS[k] for k in manual_keys],
                "Value (from the 10-K)": [saved_manual.get(k) for k in manual_keys],
            })
            edited_manual = st.data_editor(
                manual_df, key=f"manual_editor_{c['ticker']}_{c['y2_label']}", hide_index=True,
                num_rows="fixed", use_container_width=True,
                column_config={
                    "Line Item": st.column_config.TextColumn(disabled=True),
                    "Value (from the 10-K)": st.column_config.NumberColumn(format="%.2f"),
                },
            )
            manual_source_val = st.text_input(
                "Where exactly did you find these figures? (e.g. 'Apple 10-K for FY2024, filed "
                "Nov 1 2024, sec.gov')",
                value=saved.get("manual_source") or "",
                key=f"manual_source_{c['ticker']}",
            )
            manual_raw_vals = dict(zip(manual_keys, edited_manual["Value (from the 10-K)"]))
            manual_values_to_save = {k: (float(v) if pd.notna(v) else None) for k, v in manual_raw_vals.items()}
            manual_values = {k: (float(v) if pd.notna(v) else None) for k, v in manual_raw_vals.items()}
            manual_price = manual_values.pop("share_price", None)
            manual_shares = manual_values.pop("shares_outstanding", None)
            manual_market_cap = (manual_price * manual_shares) if (manual_price and manual_shares) else None
            manual_derived = compute_derived(manual_values)
            manual_ratios = compute_ratios(manual_derived, market_cap=manual_market_cap)

            st.markdown(
                "**Step 2 — copy this prompt** into your AI chat (same tab as Round 1, or a "
                "fresh one — either works):"
            )
            open_prompt = build_open_research_prompt(c)
            st.code(open_prompt, language="text")
            st.markdown("**Step 3 — paste the AI's full response here** (for your disclosure record, including its stated source/period):")
            ai_open_response_val = st.text_area(
                "Full AI response",
                value=saved.get("ai_open_response") or "",
                key=f"ai_open_response_{c['ticker']}", height=200,
                placeholder="Paste the AI's full response here, including its Data Source & Period section...",
            )
            ai_open_tool, ai_open_timestamp = record_ai_disclosure(
                "r2", c["ticker"], ai_open_response_val,
                saved_tool=saved.get("r2_tool"), saved_timestamp=saved.get("r2_stamp"),
            )
            st.markdown(
                "**Step 4 — transcribe the AI's 'Final Answers' table below**, matching each "
                "ratio to the plain number it gave you (leave blank if it answered N/A):"
            )
            saved_open_values = saved.get("ai_open_values", {})
            open_df = pd.DataFrame({
                "Ratio": [r["name"] for r in RATIO_DEFS],
                "Category": [r["category"] for r in RATIO_DEFS],
                "AI's Value": [saved_open_values.get(r["key"]) for r in RATIO_DEFS],
            })
            edited_open = st.data_editor(
                open_df, key=f"ai_open_editor_{c['ticker']}_{c['y2_label']}", hide_index=True,
                num_rows="fixed", use_container_width=True,
                column_config={
                    "Ratio": st.column_config.TextColumn(disabled=True),
                    "Category": st.column_config.TextColumn(disabled=True),
                    "AI's Value": st.column_config.NumberColumn(
                        "AI's Value (plain number — no %, x, or 'days')", format="%.4f"
                    ),
                },
            )
            rows2 = []
            for rdef, ai_val in zip(RATIO_DEFS, edited_open["AI's Value"]):
                app_raw = c["ratios_y2"][rdef["key"]]
                app_disp = display_magnitude(app_raw, rdef["suffix"])
                manual_raw_val = manual_ratios.get(rdef["key"])
                manual_disp = display_magnitude(manual_raw_val, rdef["suffix"])
                ai_disp = float(ai_val) if pd.notna(ai_val) else None
                predict_raw = predictions.get(rdef["key"])
                predict_disp = float(predict_raw) if pd.notna(predict_raw) else None
                rows2.append({
                    "Category": rdef["category"],
                    "Ratio": rdef["name"],
                    "Your Prediction": (f"{predict_disp:,.2f}" if predict_disp is not None else "N/A"),
                    "App (yfinance)": fmt(app_raw, rdef["suffix"]),
                    "Manual (10-K)": (fmt(manual_raw_val, rdef["suffix"]) if manual_raw_val is not None else "N/A"),
                    "AI (self-sourced)": (f"{ai_disp:,.2f}" if ai_disp is not None else "N/A"),
                    "You vs App": compare_flag(predict_disp, app_disp),
                    "App vs Manual": compare_flag(app_disp, manual_disp),
                    "App vs AI": compare_flag(app_disp, ai_disp),
                    # Internal-only raw values (in fmt()'s display units) for the charts below —
                    # not shown in the table itself.
                    "_app_disp": app_disp, "_manual_disp": manual_disp, "_ai_disp": ai_disp,
                    "_predict_disp": predict_disp, "_suffix": rdef["suffix"],
                })
            display_cols2 = ["Category", "Ratio", "Your Prediction", "App (yfinance)", "Manual (10-K)",
                              "AI (self-sourced)", "You vs App", "App vs Manual", "App vs AI"]
            cmp_df2 = pd.DataFrame(rows2)[display_cols2]
            st.markdown(
                "**Four-way comparison — your gut-check prediction vs. App (yfinance) vs. your "
                "Manual (10-K) figures vs. the AI's self-sourced calculation:**"
            )
            st.dataframe(cmp_df2, use_container_width=True, hide_index=True)
            n_match_you = sum(1 for r in rows2 if r["You vs App"] == "✅ Match")
            n_diff_you = sum(1 for r in rows2 if r["You vs App"] == "❌ Differ")
            n_match_am = sum(1 for r in rows2 if r["App vs Manual"] == "✅ Match")
            n_diff_am = sum(1 for r in rows2 if r["App vs Manual"] == "❌ Differ")
            n_match_aai = sum(1 for r in rows2 if r["App vs AI"] == "✅ Match")
            n_diff_aai = sum(1 for r in rows2 if r["App vs AI"] == "❌ Differ")
            st.caption(
                f"You vs App: ✅ {n_match_you} match · ❌ {n_diff_you} differ  |  "
                f"App vs Manual: ✅ {n_match_am} match · ❌ {n_diff_am} differ  |  "
                f"App vs AI: ✅ {n_match_aai} match · ❌ {n_diff_aai} differ"
            )
            SUFFIX_LABEL = {"": "plain ratio", "%": "%", "x": "×", " days": "days"}
            show_charts2 = st.checkbox(
                "📊 Show per-ratio comparison charts (You vs App vs Manual vs AI)",
                key=f"show_charts_open_{c['ticker']}",
            )
            if show_charts2:
                chart_cols = st.columns(3)
                for i, row in enumerate(rows2):
                    with chart_cols[i % 3]:
                        bars = {}
                        if row["_predict_disp"] is not None:
                            bars["You"] = row["_predict_disp"]
                        if row["_app_disp"] is not None:
                            bars["App"] = row["_app_disp"]
                        if row["_manual_disp"] is not None:
                            bars["Manual"] = row["_manual_disp"]
                        if row["_ai_disp"] is not None:
                            bars["AI"] = row["_ai_disp"]
                        st.caption(f"**{row['Ratio']}** ({SUFFIX_LABEL.get(row['_suffix'], '')})")
                        if bars:
                            st.bar_chart(pd.DataFrame({"Value": bars}), height=180)
                        else:
                            st.caption("No values entered yet.")
            st.markdown("**Your reflection on Round 2** — required:")
            reflection_val = st.text_area(
                "How did your own prediction compare to the App, Manual, and AI values? Where do "
                "those three differ from each other, and why? Is it a data-source difference "
                "(App's yfinance data vs. your 10-K figures — e.g. different fiscal period, "
                "restated numbers, mislabeled line items), a research/calculation difference on "
                "the AI's part, or both? Which would you trust most for a real assignment, and why?",
                value=saved.get("open_reflection") or "",
                key=f"open_reflection_{c['ticker']}", height=140,
            )

            autosave_progress(student_username, ticker, y2, {
                "ai_calc_response": ai_calc_response_val,
                "r1_tool": ai_calc_tool, "r1_stamp": ai_calc_timestamp,
                "ai_calc_values": {r["key"]: (float(v) if pd.notna(v) else None)
                                    for r, v in zip(RATIO_DEFS, edited_calc["AI's Value"])},
                "ai_calc_reflection": st.session_state.get(f"ai_calc_reflection_{c['ticker']}", ""),
                "predictions": {k: (float(v) if pd.notna(v) else None) for k, v in predictions.items()},
                "predict_confirmed": st.session_state.get(predict_lock_key, False),
                "manual_values": manual_values_to_save,
                "manual_source": manual_source_val,
                "ai_open_response": ai_open_response_val,
                "r2_tool": ai_open_tool, "r2_stamp": ai_open_timestamp,
                "ai_open_values": {r["key"]: (float(v) if pd.notna(v) else None)
                                    for r, v in zip(RATIO_DEFS, edited_open["AI's Value"])},
                "open_reflection": reflection_val,
            })
            saved_at = st.session_state.get(f"_prog_saved_at_{ticker}_{y2}")
            if progress_store_available():
                if saved_at:
                    st.caption(f"💾 Progress autosaved at {saved_at} — you can log out and resume later.")
                else:
                    st.caption(
                        "⚠️ Autosave is set up but the last save attempt failed — your answers "
                        "are still here in this browser tab, so download your CSV/PDF as a backup."
                    )

            st.markdown("**Download this comparison** — for documentation and submission:")
            dl_r2_1, dl_r2_2 = st.columns(2)
            with dl_r2_1:
                csv_bytes2 = cmp_df2.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "CSV (table only)", data=csv_bytes2,
                    file_name=f"{c['ticker']}_round2_comparison.csv", mime="text/csv",
                    key=f"dl_r2_csv_{c['ticker']}", use_container_width=True,
                )
            with dl_r2_2:
                pdf_buf2 = build_calc_challenge_round2_pdf(
                    company=c, manual_source=manual_source_val, ai_prompt=open_prompt,
                    ai_response=ai_open_response_val, rows2=rows2, reflection=reflection_val,
                    prepared_by=prepared_by_display, ai_tool=ai_open_tool, ai_timestamp=ai_open_timestamp,
                )
                st.download_button(
                    "PDF (full report — table + prompt + AI response + reflection)",
                    data=pdf_buf2, file_name=f"{c['ticker']}_round2_report.pdf",
                    mime="application/pdf", key=f"dl_r2_pdf_{c['ticker']}", use_container_width=True,
                )
    st.caption(
        "Reminder: a 5% tolerance is used to call two values a 'Match' (rounding/format "
        "differences are normal); anything outside that is flagged for you to investigate in "
        "your reflections above. Industry benchmark values elsewhere in this app are illustrative "
        "educational reference points, not live data — always cross-check unusual figures against "
        "the company's actual financial filings."
    )
